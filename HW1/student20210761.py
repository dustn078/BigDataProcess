import openpyxl

file_path = "student.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

def calculate_grade(total_score):
    if total_score < 40:
        return "F"
    elif total_score >= 90:
        return "A+"
    elif total_score >= 85:
        return "A0"
    elif total_score >= 80:
        return "B+"
    elif total_score >= 75:
        return "B0"
    elif total_score >= 70:
        return "C+"
    else:
        return "C0"


for row in sheet.iter_rows(min_row=2, values_only=True):
    student_id, name, midterm, final, hw, attendance = row[0], row[1], row[2], row[3], row[4], row[5]
    total_score = midterm * 0.3 + final * 0.35 + hw * 0.34 + attendance * 0.01
    grade = calculate_grade(total_score)

    sheet.cell(row=student_id, column=7, value=total_score)
    sheet.cell(row=student_id, column=8, value=grade)

workbook.save(file_path)
workbook.close()